##Python program for reading and writing into excel using openpyxl downloaded from pypi
import openpyxl

# Inventory file contains Product No, Inventory, Price Supplier
inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inventory_file["Sheet1"]

# Dictionary to hold supplier names and their product count
products_per_supplier = {}
inventory_value_per_supplier = {}
products_under_10_inv = {}

print(f"Product Size is {product_list.max_row}")

""" Problem 1: Count and display number of supplier and their product count"""
""" Problem 2: Total inventory value per each supplier """
""" Problem 3: List products under 10 for each supplier"""
""" Problem 3: Add a new column and update with the price of each product inventory """
for product_row in range(2, product_list.max_row + 1):
        supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)



    #Problem 1
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        products_per_supplier[supplier_name] = 1

    #Problem 2
    if supplier_name in inventory_value_per_supplier:
        inventory_value_per_supplier[supplier_name] = inventory_value_per_supplier.get(supplier_name) + inventory * price
    else:
        inventory_value_per_supplier[supplier_name] = inventory * price

    #Problem 3
    if inventory< 10 :
        products_under_10_inv[int(product_num)] = int(inventory)

    #Problem 4
    inventory_price.value = inventory * price

print(products_per_supplier)
print(inventory_value_per_supplier)
print(products_under_10_inv)

inventory_file.save("myInventorySolution.xlsx")